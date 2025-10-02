import os, zipfile, mimetypes, traceback
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Dashboard Doanh Thu", layout="wide")

# =========================================================
# HÀM DEBUG & ĐỌC FILE AN TOÀN
# =========================================================
def read_file_debug(path, label="file"):
    st.markdown(f"### 🔍 Debug đọc {label}: `{path}`")

    if not os.path.exists(path):
        st.error(f"❌ Không tìm thấy file `{path}` trong thư mục hiện tại.")
        return None

    try:
        size = os.path.getsize(path)
        st.write(f"✅ File tồn tại — kích thước: {size:,} bytes")
    except Exception as e:
        st.warning(f"Không lấy được kích thước file: {e}")

    mt = mimetypes.guess_type(path)
    st.write("Mime/type dự đoán:", mt)

    # Nếu là xlsx thì check zip
    try:
        if zipfile.is_zipfile(path):
            with zipfile.ZipFile(path) as z:
                st.write("Là file ZIP (xlsx) — chứa các sheet:", z.namelist()[:10])
        else:
            st.write("Không phải file ZIP hợp lệ (.xlsx có thể bị lỗi).")
    except Exception as e:
        st.warning(f"Zip inspect error: {e}")

    # Thử mở bằng openpyxl
    try:
        wb = load_workbook(path, read_only=True)
        sheets = wb.sheetnames
        st.success(f"openpyxl: mở được — sheets: {sheets}")
        preview = []
        for i, row in enumerate(wb[sheets[0]].iter_rows(values_only=True, max_row=5)):
            preview.append(row)
        st.write("Preview sheet đầu:", preview)
        wb.close()
    except Exception:
        st.error("openpyxl không đọc được file.")
        st.text(traceback.format_exc())

    # Thử pandas.read_excel
    try:
        df = pd.read_excel(path, engine="openpyxl")
        st.success(f"pandas.read_excel thành công — shape: {df.shape}")
        return df
    except Exception:
        st.error("pandas.read_excel thất bại.")
        st.text(traceback.format_exc())

    # Thử pandas.read_csv fallback
    try:
        df2 = pd.read_csv(path)
        st.success(f"pandas.read_csv thành công — shape: {df2.shape}")
        return df2
    except Exception:
        st.error("pandas.read_csv cũng thất bại.")
        st.text(traceback.format_exc())

    return None


# =========================================================
# ĐỌC FILE (CÓ UPLOAD NẾU LỖI)
# =========================================================
file_now = "data.xlsx"    # tháng hiện tại
file_old = "data3.xlsx"   # 3 tháng trước
file_map = "nh.xlsx"      # mapping ngành hàng -> nhóm

df_now = read_file_debug(file_now, "data (tháng hiện tại)")
df_old = read_file_debug(file_old, "data3 (3 tháng trước)")
df_map = read_file_debug(file_map, "mapping (nh.xlsx)")

# Nếu có file nào không đọc được thì cho upload tay
if df_now is None or df_old is None or df_map is None:
    st.warning("⚠️ Một hoặc nhiều file chưa đọc được. Vui lòng upload tay.")

    up_now = st.file_uploader("Upload data.xlsx", type=["xlsx", "csv"])
    up_old = st.file_uploader("Upload data3.xlsx", type=["xlsx", "csv"])
    up_map = st.file_uploader("Upload nh.xlsx", type=["xlsx", "csv"])

    if up_now:
        try:
            df_now = pd.read_excel(up_now, engine="openpyxl")
        except Exception:
            up_now.seek(0)
            df_now = pd.read_csv(up_now)

    if up_old:
        try:
            df_old = pd.read_excel(up_old, engine="openpyxl")
        except Exception:
            up_old.seek(0)
            df_old = pd.read_csv(up_old)

    if up_map:
        try:
            df_map = pd.read_excel(up_map, engine="openpyxl")
        except Exception:
            up_map.seek(0)
            df_map = pd.read_csv(up_map)

    if df_now is None or df_old is None or df_map is None:
        st.stop()  # Dừng app cho tới khi có file hợp lệ

# =========================================================
# TIỀN XỬ LÝ DỮ LIỆU
# =========================================================
st.subheader("📊 Tổng quan dữ liệu")

st.write("Data tháng hiện tại:", df_now.shape)
st.dataframe(df_now.head())

st.write("Data 3 tháng trước:", df_old.shape)
st.dataframe(df_old.head())

st.write("Mapping ngành hàng:", df_map.shape)
st.dataframe(df_map.head())

# Merge mapping để lấy nhóm ngành
if "Ngành hàng" in df_now.columns and "Ngành hàng" in df_map.columns:
    df_now = df_now.merge(df_map, on="Ngành hàng", how="left")
    df_old = df_old.merge(df_map, on="Ngành hàng", how="left")

# =========================================================
# VÍ DỤ: CHỈ SỐ DOANH THU THEO NHÓM
# =========================================================
if "Nhóm" in df_now.columns and "Tổng doanh thu" in df_now.columns:
    st.subheader("📈 Doanh thu theo nhóm ngành (tháng hiện tại)")
    st.bar_chart(df_now.groupby("Nhóm")["Tổng doanh thu"].sum())
